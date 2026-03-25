#!/usr/bin/env python3
"""
Применение всех исправлений из Regon.xlsx в region_assignments.js
и пересборка регионов в nations.js.
"""
import openpyxl, re, json

def parse_js_obj(text, varname):
    m = re.search(r'(?:var|const|let)\s+' + varname + r'\s*=\s*(\{)', text)
    if not m: return None
    start = m.start(1)
    depth = 0
    for i in range(start, len(text)):
        if text[i] == '{': depth += 1
        elif text[i] == '}':
            depth -= 1
            if depth == 0:
                return json.loads(text[start:i+1])
    return None

# ── Маппинг: xlsx страна → game nation_id ────────────────────────────
NAME_MAP = {
    'Etruscan Confederation': 'etruscan_conf', 'Roman Republic': 'rome',
    'Syracuse': 'syracuse', 'Carthage': 'carthage', 'Lucani': 'lucani',
    'Samnites': 'samnites', 'Massilia': 'massilia', 'Apuania': 'apuania',
    'Apulians': 'apulians', 'Boi': 'boii', 'Brutti': 'brutii',
    'Camuni': 'camuni', 'Caraceni': 'caraceni', 'Carenses': 'carenses',
    'Caturigi': 'caturigi', 'Cenomanni': 'cenomanni', 'Croton': 'croton',
    'Elea': 'elea', 'Frentani': 'frentani', 'Gela': 'gela', 'Genua': 'genua',
    'Herakleia': 'herakleia', 'Ingaunia': 'ingaunia', 'Insubri': 'insubri',
    'Laevi': 'laevi', 'Lingones': 'lingones', 'Locri': 'locri', 'Marsi': 'marsi',
    'Marrucini': 'marrucini', 'Messapains': 'messapians', 'Metapontum': 'metapontum',
    'Paeligni': 'paeligni', 'Picentes': 'picentes', 'Rugusci': 'rugusci',
    'Salassi': 'salassi', 'Senones': 'senones', 'Sicels': 'sicels', 'Spina': 'spina',
    'Statiallia': 'statiellia', 'Taras': 'taras', 'Taurini': 'taurini',
    'Tridentini': 'tridentini', 'Umbrians': 'umbrians', 'Veneti': 'venetii',
    'Acragas': 'acragas', 'Ambisonti': 'ambisonti', 'Leponti': 'leponti',
    'Anauni': 'anauni', 'Ancona': 'ancona', 'Brundisium': 'brundisium',
    'Calactea': 'calactea', 'Celesitani': 'celesitani', 'Coracenses': 'coracenses',
    'Docilia': 'docilia', 'Elymia': 'elymia', 'Feltrini': 'feltrini',
    'Graioceli': 'graioceli', 'Herakleia Minoa': 'herakleia_minoa',
    'Hipponion': 'hipponion', 'Ilvatia': 'ilvatia', 'Intimilia': 'intimilia',
    'Isarci': 'isarci', 'Kallipolis': 'kallipolis', 'Lai': 'lai',
    'Maricia': 'maricia', 'Nuceria': 'nuceria', 'Ravenna': 'ravenna',
    'Rhegium': 'rhegium', 'Sabinia': 'sabinia', 'Selinous': 'selinous',
    'Sicani': 'sicani', 'Sipontum': 'sipontum', 'Thuria': 'thuria',
    'Turritani': 'turritani', 'Tyndaria': 'tyndaria', 'Uberi': 'uberi',
    'Vediantia': 'vediantia', 'Vertamocorii': 'vertamocorii', 'Vestini': 'vestini',
    # Новые нации:
    'Praetutil': 'praetuttii', 'Caluconi': 'caluconi', 'Suaneti': 'suaneti',
    'Anamaria': 'anamaria', 'Lapicinia': 'lapicinia', 'Libuia': 'libuia',
    'Orobia': 'orobii', 'Tigullia': 'tigullia', 'Dectuninia': 'dectuninia',
    'Vagiennia': 'vagiennia', 'Ilvatia Coeba': 'ilvatia_coeba',
    'Iadatinia': 'iadatinia', 'Votodronia': 'votodronia', 'Oxybia': 'oxybia',
    'Vesubia': 'vesubia', 'Varagria': 'varagria', 'Seduni': 'seduni',
    'Trumpili': 'trumpili', 'Venosti': 'venosti', 'Breuni': 'breuni',
    'Sevaci': 'sevaci', 'Ambidravi': 'ambidravi', 'Genauni': 'genauni',
    'Focunati': 'focunati', 'Rubrensians': 'rubrensians', 'Valentini': 'valentini',
    'Luquidonenses': 'luquidonenses', 'Tibulati': 'tibulati',
    'Longonenses': 'longonenses', 'Maricia': 'maricia',
}

# ── Читаем xlsx ───────────────────────────────────────────────────────
wb = openpyxl.load_workbook('Regon.xlsx')
ws = wb.active

xlsx_entries = {}
for r in range(3, ws.max_row + 1):
    rid = ws.cell(r, 1).value
    name = ws.cell(r, 2).value
    country = ws.cell(r, 3).value
    if rid and name and country:
        xlsx_entries[str(rid).strip()] = (str(name).strip(), str(country).strip())

# ── Читаем текущие assignments ────────────────────────────────────────
with open('data/region_assignments.js') as f:
    atext = f.read()
assignments = parse_js_obj(atext, 'REGION_ASSIGNMENTS')

# ── Применяем изменения ───────────────────────────────────────────────
changes = 0
skipped_unknown = []

for rid, (region_name, country_name) in xlsx_entries.items():
    nation_id = NAME_MAP.get(country_name)
    if not nation_id:
        nation_id = country_name.lower().replace(' ', '_')

    current = assignments.get(rid, '?')
    if current == nation_id:
        continue  # уже верно

    # Apply change
    new_text = re.sub(
        r'("' + re.escape(rid) + r'"\s*:\s*)"[^"]*"',
        r'\1"' + nation_id + '"',
        atext
    )
    if new_text != atext:
        changes += 1
        atext = new_text
        print(f'  {rid:8} {current:<22} → {nation_id:<22}  [{country_name}]')
    else:
        print(f'  WARN {rid} not found in assignments!')

with open('data/region_assignments.js', 'w') as f:
    f.write(atext)

print(f'\n✓ Изменено {changes} регионов в region_assignments.js')

# ── Пересобираем regions в nations.js ────────────────────────────────
print('\nПересобираем regions в nations.js...')

# Re-read assignments after changes
assignments_new = parse_js_obj(atext, 'REGION_ASSIGNMENTS')

# Build nation → regions index
nation_to_regions = {}
for rid, nid in assignments_new.items():
    if nid in ('neutral',) or not rid.startswith('r'):
        continue
    if nid not in nation_to_regions:
        nation_to_regions[nid] = []
    nation_to_regions[nid].append(rid)

# Sort each list
for nid in nation_to_regions:
    nation_to_regions[nid].sort(key=lambda x: int(x[1:]))

# Update nations.js
with open('data/nations.js') as f:
    ntext = f.read()

nations_changed = 0
for nid, new_regions in sorted(nation_to_regions.items()):
    new_regions_sorted = sorted(set(new_regions), key=lambda x: int(x[1:]))

    # Try to update the regions array for this nation
    pattern = r'("' + re.escape(nid) + r'"(?:[^{}]|\{[^{}]*\})*?"regions"\s*:\s*)\[[^\]]*\]'
    def replacer(m):
        return m.group(1) + json.dumps(new_regions_sorted)
    new_ntext = re.sub(pattern, replacer, ntext, count=1)
    if new_ntext != ntext:
        ntext = new_ntext
        nations_changed += 1

with open('data/nations.js', 'w') as f:
    f.write(ntext)

print(f'✓ Обновлено {nations_changed} наций в nations.js')
